from flask import Flask, render_template, request, jsonify, send_file, Response
import os
import sys
import tempfile
import base64
import json
from io import BytesIO
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Import modules lazily to avoid startup failures
try:
    from textract_processor import extract_text_from_pdf, extract_text_from_pdf_bytes, extract_structured_data_from_pdf_bytes
    from llm_processor import process_text_with_llm
    from structured_llm_processor import process_structured_data_with_llm, process_structured_data_with_llm_unified
    from export_utils import export_to_pdf
    MODULES_LOADED = True
    print("✅ All modules loaded successfully")
except ImportError as e:
    print(f"⚠️ Module import error: {e}")
    MODULES_LOADED = False

# Configurable flag for unified context extraction
USE_UNIFIED_CONTEXT_EXTRACTION = True  # Set False to keep legacy context tracker

app = Flask(__name__)

# Ensure templates directory exists
os.makedirs('templates', exist_ok=True)
os.makedirs('static', exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health_check():
    """Simple health check endpoint for Railway"""
    try:
        return jsonify({
            'status': 'healthy',
            'message': 'PDF Extraction Service is running',
            'modules_loaded': MODULES_LOADED
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/status')
def detailed_status():
    """Detailed status check with module verification"""
    import shutil
    
    status = {
        'status': 'healthy',
        'tesseract_available': bool(shutil.which('tesseract')),
        'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        'modules_loaded': MODULES_LOADED
    }
    
    # Check if we can import required modules
    if MODULES_LOADED:
        try:
            import openai, boto3, pandas, fitz
            status['modules_ok'] = True
        except ImportError as e:
            status['modules_ok'] = False
            status['import_error'] = str(e)
    else:
        status['modules_ok'] = False
        status['import_error'] = 'Modules not loaded at startup'
    
    return jsonify(status)

@app.route('/extract', methods=['POST'])
def extract():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        # Extract structured data from PDF using Amazon Textract
        pdf_bytes = file.read()
        structured_data = extract_structured_data_from_pdf_bytes(pdf_bytes)
        
        # Return the new JSON format
        return jsonify(structured_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def summarize_commentary(text):
    """Summarize long commentary using GPT-4o"""
    try:
        import openai
        import os
        
        client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
        
        prompt = f"""Summarize this financial document commentary in 2-3 complete sentences, preserving all key information:

{text}

Instructions:
- Preserve ALL financial figures, percentages, dates, and company names
- Keep the complete meaning and context
- Use complete sentences that don't cut off mid-thought
- Maintain the professional tone and key details"""

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=150,
            temperature=0.2
        )
        
        # Calculate and log cost for summarization
        if hasattr(response, 'usage') and response.usage:
            input_tokens = response.usage.prompt_tokens
            output_tokens = response.usage.completion_tokens
            input_cost = (input_tokens / 1_000_000) * 5.00  # GPT-4o input cost
            output_cost = (output_tokens / 1_000_000) * 15.00  # GPT-4o output cost
            total_cost = input_cost + output_cost
            print(f"Commentary summarization cost: ${total_cost:.6f} ({input_tokens} input + {output_tokens} output tokens)")
        
        return response.choices[0].message.content.strip() if response.choices[0].message.content else ""
    except Exception as e:
        print(f"Error summarizing commentary: {e}")
        return text[:200] + '...' if len(text) > 200 else text

def convert_unified_to_standard_format(unified_result: dict, original_data: dict) -> dict:
    """
    Convert unified extraction result to standard format for compatibility.
    
    Args:
        unified_result: Result from process_structured_data_with_llm_unified
        original_data: Original structured data for fallback
        
    Returns:
        Result in standard format with enhanced_data_with_context
    """
    enhanced_data = []
    
    if "fields" in unified_result:
        for field_name, field_data in unified_result["fields"].items():
            if isinstance(field_data, dict):
                enhanced_data.append({
                    'source': 'Unified Extraction',
                    'type': 'Financial Data',
                    'field': field_name,
                    'value': field_data.get('value', ''),
                    'page': 'N/A',
                    'context': field_data.get('context', ''),
                    'has_context': bool(field_data.get('context', '').strip())
                })
    
    # Add cost summary from global tracker
    from structured_llm_processor import cost_tracker
    cost_summary = cost_tracker.get_summary()
    
    return {
        'enhanced_data_with_context': enhanced_data,
        'cost_summary': cost_summary,
        'processing_mode': 'unified_extraction'
    }


def clean_csv_value(value):
    """Clean value for CSV format - escape commas and quotes"""
    if not value:
        return ""
    
    # Convert to string and clean up
    value_str = str(value).strip()
    
    # Escape quotes by doubling them
    value_str = value_str.replace('"', '""')
    
    # If value contains comma, newline, or quote, wrap in quotes
    if ',' in value_str or '\n' in value_str or '"' in value_str:
        value_str = f'"{value_str}"'
    
    return value_str

def find_relevant_document_context(field, value, document_text):
    """Find relevant context from document text for a specific field/value pair"""
    if not document_text or not field:
        return ""
    
    field_lower = field.lower().replace('_', ' ')
    value_lower = str(value).lower().replace('$', '').replace('%', '').replace(',', '')
    
    # Extract numeric part if value contains numbers
    import re
    numeric_parts = re.findall(r'\d+\.?\d*', value_lower)
    
    best_matches = []
    
    # Look for text segments that mention the field or value
    for i, line in enumerate(document_text):
        line_lower = line.lower()
        score = 0
        
        # High priority: exact value match
        if value_lower and len(value_lower) > 2 and value_lower in line_lower:
            score += 10
        
        # Medium priority: numeric match
        for num in numeric_parts:
            if len(num) > 1 and num in line_lower:
                score += 5
        
        # Lower priority: field word matches
        field_words = [word for word in field_lower.split() if len(word) > 2]
        for word in field_words:
            if word in line_lower:
                score += 2
        
        # If we found a relevant match, store it with context
        if score >= 7:  # Only include high-confidence matches
            # Get context around the matching line
            start_idx = max(0, i - 1)
            end_idx = min(len(document_text), i + 2)
            context_lines = document_text[start_idx:end_idx]
            
            # Join and clean up the context
            context = ' '.join(context_lines).strip()
            
            best_matches.append({
                'text': context,
                'score': score,
                'line_index': i
            })
    
    # Sort by score and return the best match, truncated if too long
    if best_matches:
        best_matches.sort(key=lambda x: x['score'], reverse=True)
        best_context = best_matches[0]['text']
        
        # Truncate if too long but keep complete sentences
        if len(best_context) > 300:
            sentences = best_context.replace('!', '.').replace('?', '.').split('.')
            complete_text = ""
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence and len(complete_text + sentence) < 250:
                    complete_text += sentence + ". "
                else:
                    break
            
            return complete_text.strip() if complete_text else best_context[:300] + '...'
        else:
            return best_context
    
    return ""

def find_relevant_document_text(row_data, document_text):
    """Find relevant text from document that mentions this data point"""
    field = row_data.get('field', '').lower()
    value = str(row_data.get('value', '')).lower()
    
    # Clean field and value for better matching
    field_words = [word for word in field.replace('_', ' ').split() if len(word) > 2]
    value_clean = value.replace('$', '').replace('%', '').replace(',', '').strip()
    
    # Extract numeric part if value contains numbers
    import re
    numeric_part = re.findall(r'\d+\.?\d*', value_clean)
    
    best_matches = []
    
    # Look for text segments that mention the value or field
    for i, line in enumerate(document_text):
        line_lower = line.lower()
        line_clean = _clean_superscript_numbers(line_lower)
        score = 0
        
        # High priority: exact value match
        if value_clean and len(value_clean) > 2 and value_clean in line_clean:
            score += 10
        
        # Medium priority: numeric match
        for num in numeric_part:
            if len(num) > 1 and num in line_clean:
                score += 7
        
        # Lower priority: field word matches
        for word in field_words:
            if word in line_lower:
                score += 2
        
        # If we found a relevant match, store it with context
        if score >= 9:  # Only include very high-confidence matches
            # Get targeted context around the matching line
            start_idx = max(0, i - 1)
            end_idx = min(len(document_text), i + 3)
            context_lines = document_text[start_idx:end_idx]
            
            # Join and clean up the context
            context = ' '.join(context_lines).strip()
            context = _clean_superscript_numbers(context)
            
            best_matches.append({
                'text': context,
                'score': score,
                'line_index': i
            })
    
    # Sort by score and return the best match
    if best_matches:
        best_matches.sort(key=lambda x: x['score'], reverse=True)
        best_context = best_matches[0]['text']
        
        # Truncate if too long but keep complete sentences
        if len(best_context) > 400:
            sentences = best_context.replace('!', '.').replace('?', '.').split('.')
            complete_text = ""
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence and len(complete_text + sentence) < 350:
                    complete_text += sentence + ". "
                else:
                    break
            
            if complete_text:
                return complete_text.strip()
            else:
                return best_context[:400] + '...'
        else:
            return best_context
    
    return ''  # No relevant matches found

def _clean_superscript_numbers(text):
    """Remove superscript numbers from text for better matching"""
    import re
    
    # Remove superscript numbers (Unicode superscript characters)
    superscript_pattern = r'[⁰¹²³⁴⁵⁶⁷⁸⁹]+'
    text = re.sub(superscript_pattern, '', text)
    
    # Remove common footnote reference patterns
    footnote_patterns = [
        r'\(\d+\)',    # (1), (2), etc.
        r'\[\d+\]',    # [1], [2], etc.
        r'\*+',        # *, **, ***, etc.
    ]
    
    for pattern in footnote_patterns:
        text = re.sub(pattern, '', text)
    
    return ' '.join(text.split())

def get_unmatched_document_text(df_data, document_text):
    """Get document text that doesn't match any extracted data"""
    used_indices = set()
    
    # Mark lines that were used for commentary (with context)
    for row in df_data:
        if row.get('commentary'):
            commentary_sample = row['commentary'][:100].lower()
            for i, line in enumerate(document_text):
                if commentary_sample in line.lower():
                    # Mark this line and surrounding context as used
                    for j in range(max(0, i-1), min(len(document_text), i+2)):
                        used_indices.add(j)
    
    # Collect unused lines in meaningful paragraphs
    unmatched_paragraphs = []
    current_paragraph = []
    
    for i, line in enumerate(document_text):
        if i not in used_indices and len(line.strip()) > 15:
            current_paragraph.append(line.strip())
        else:
            # End of paragraph - save if substantial
            if current_paragraph:
                paragraph_text = ' '.join(current_paragraph)
                if len(paragraph_text) > 50:  # Only keep substantial paragraphs
                    unmatched_paragraphs.append(paragraph_text)
                current_paragraph = []
    
    # Don't forget the last paragraph
    if current_paragraph:
        paragraph_text = ' '.join(current_paragraph)
        if len(paragraph_text) > 50:
            unmatched_paragraphs.append(paragraph_text)
    
    # Limit and truncate paragraphs for readability with complete sentences
    final_chunks = []
    for paragraph in unmatched_paragraphs[:3]:  # Limit to 3 substantial chunks
        if len(paragraph) > 500:
            # Find complete sentences to avoid cutting off mid-sentence
            sentences = paragraph.replace('!', '.').replace('?', '.').split('.')
            complete_paragraph = ""
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence and len(complete_paragraph + sentence) < 450:
                    complete_paragraph += sentence + ". "
                else:
                    break
            
            if complete_paragraph and len(complete_paragraph) > 50:
                final_chunks.append(complete_paragraph.strip())
            else:
                # Fallback: truncate at word boundary
                truncated = paragraph[:450]
                last_space = truncated.rfind(' ')
                if last_space > 300:
                    final_chunks.append(truncated[:last_space] + '...')
        else:
            final_chunks.append(paragraph)
    
    return final_chunks

@app.route('/process_stream', methods=['POST'])
def process_stream():
    """Streaming endpoint for progressive data processing in XLSX format"""
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    def generate():
        try:
            # Process the structured JSON data using unified or legacy mode
            if USE_UNIFIED_CONTEXT_EXTRACTION:
                # Pass full structured data for comprehensive unified processing
                unified_result = process_structured_data_with_llm_unified(data)
                
                # Convert unified result to standard format
                result = convert_unified_to_standard_format(unified_result, data)
            else:
                # Use legacy processing
                result = process_structured_data_with_llm(data)
            
            # Initialize CSV-like output for XLSX format
            csv_output = []
            row_counter = 1
            commentary_collection = []  # Collect all commentary for single row
            
            # Start with header row
            header_row = f"row{row_counter}: source,type,field,value,page,context"
            csv_output.append(header_row)
            yield f"data: {json.dumps({'type': 'header', 'content': header_row})}\n\n"
            row_counter += 1
            
            # Use enhanced data with context if available
            if 'enhanced_data_with_context' in result and result['enhanced_data_with_context']:
                print(f"Using enhanced data with context: {len(result['enhanced_data_with_context'])} rows")
                for row_data in result['enhanced_data_with_context']:
                    context = row_data.get('context', '')
                    csv_row = f"row{row_counter}: {row_data.get('source', '')},{row_data.get('type', '')},{row_data.get('field', '')},{clean_csv_value(str(row_data.get('value', '')))},{row_data.get('page', 'N/A')},{clean_csv_value(context)}"
                    csv_output.append(csv_row)
                    yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                    row_counter += 1
            elif USE_UNIFIED_CONTEXT_EXTRACTION:
                # In unified mode with no extracted data, add a message explaining why
                print("Unified mode extracted no data - likely insufficient content")
                message_row = f"row{row_counter}: Unified Extraction,Information,No Data Extracted,No meaningful data found in document,N/A,The document appears to contain insufficient content for data extraction. Please ensure the document contains financial data tables or structured information."
                csv_output.append(message_row)
                yield f"data: {json.dumps({'type': 'row', 'content': message_row})}\n\n"
                row_counter += 1
            else:
                # Fallback to original processing if enhanced data is not available
                # Process tables with enhanced financial data extraction
                if 'processed_tables' in result and result['processed_tables']:
                    for i, table in enumerate(result['processed_tables']):
                        if table.get('structured_table') and not table['structured_table'].get('error'):
                            table_data = table['structured_table']
                            page = table.get('page', 'N/A')
                            
                            # Stream each table data row in XLSX format (no individual context)
                            for key, value in table_data.items():
                                if key != 'error' and value:
                                    # Create CSV-formatted row without individual context
                                    csv_row = f"row{row_counter}: Table {i+1},Table Data,{key},{clean_csv_value(str(value))},{page},"
                                    csv_output.append(csv_row)
                                    yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                                    row_counter += 1
            
                # Process key-value pairs (fallback)
                if 'processed_key_values' in result and result['processed_key_values']:
                    kv_data = result['processed_key_values'].get('structured_key_values', {})
                    if kv_data and not kv_data.get('error'):
                        for key, value in kv_data.items():
                            if key != 'error' and value:
                                # Create CSV-formatted row without individual context
                                csv_row = f"row{row_counter}: Key-Value Pairs,Structured Data,{key},{clean_csv_value(str(value))},N/A,"
                                csv_output.append(csv_row)
                                yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                                row_counter += 1
                
                # Process document text facts and collect commentary (fallback)
                if 'processed_document_text' in result and result['processed_document_text']:
                    for chunk_idx, chunk in enumerate(result['processed_document_text']):
                        if 'extracted_facts' in chunk and not chunk['extracted_facts'].get('error'):
                            facts = chunk['extracted_facts']
                            for key, value in facts.items():
                                if key != 'error' and value:
                                    # Determine if this is footnote content
                                    data_type = 'Footnote' if 'footnote' in key.lower() else 'Financial Data'
                                    field_name = key.replace('_Footnote', ' (Footnote)').replace('Footnote_', 'Footnote: ')
                                    
                                    # Create CSV-formatted row without individual context
                                    csv_row = f"row{row_counter}: Text Chunk {chunk_idx+1},{data_type},{field_name},{clean_csv_value(str(value))},N/A,"
                                    csv_output.append(csv_row)
                                    yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                                    row_counter += 1
                                    
                                    # Collect for general commentary
                                    commentary_collection.append(f"{field_name}: {value}")
            
            # Collect document text for general commentary
            document_text_list = data.get('document_text', [])
            if document_text_list:
                # Clean and organize document text into paragraphs
                text_content = []
                for i, line in enumerate(document_text_list):
                    line = line.strip()
                    if line and len(line) > 20:  # Only include substantial lines
                        text_content.append(line)
                
                # Group into meaningful paragraphs
                if text_content:
                    commentary_collection.append("\n\nDocument Overview:")
                    paragraph_break_counter = 0
                    for line in text_content[:10]:  # Limit to first 10 substantial lines
                        commentary_collection.append(line)
                        paragraph_break_counter += 1
                        if paragraph_break_counter % 3 == 0:  # Add paragraph break every 3 lines
                            commentary_collection.append("")
            
            # Process standalone footnotes and collect for commentary
            if 'footnotes' in data and data['footnotes']:
                footnote_texts = []
                for footnote in data['footnotes']:
                    footnote_content = footnote.get('content', '')
                    marker = footnote.get('marker', 'N/A')
                    footnote_texts.append(f"Footnote {marker}: {footnote_content}")
                
                if footnote_texts:
                    commentary_collection.append("\n\nFootnotes:")
                    commentary_collection.extend(footnote_texts)
            
            # Create comprehensive general commentary row
            all_commentary = []
            
            # Add extracted data summary
            if commentary_collection:
                all_commentary.extend(commentary_collection)
            
            # Add comprehensive document text analysis
            document_text_list = data.get('document_text', [])
            if document_text_list:
                all_commentary.append("Full Document Analysis:")
                
                # Process all substantial text lines
                substantial_lines = []
                for line in document_text_list:
                    cleaned_line = line.strip()
                    if cleaned_line and len(cleaned_line) > 25:  # Only substantial content
                        substantial_lines.append(cleaned_line)
                
                # Group lines into coherent paragraphs
                if substantial_lines:
                    current_paragraph = []
                    for i, line in enumerate(substantial_lines[:20]):  # Limit to first 20 substantial lines
                        current_paragraph.append(line)
                        
                        # Create paragraph breaks every 3-4 lines or at natural breaks
                        if len(current_paragraph) >= 3 or i == len(substantial_lines) - 1:
                            paragraph_text = " ".join(current_paragraph)
                            if len(paragraph_text) > 50:  # Only include substantial paragraphs
                                all_commentary.append(paragraph_text)
                            current_paragraph = []
            
            # Add cost and processing summary
            cost_summary = result.get('cost_summary', {})
            if cost_summary:
                total_cost = cost_summary.get('total_cost_usd', 0)
                total_tokens = cost_summary.get('total_tokens', 0)
                api_calls = cost_summary.get('api_calls', 0)
                all_commentary.append(f"Processing Summary: This analysis used ${total_cost:.6f} in AI processing costs, consuming {total_tokens:,} tokens across {api_calls} API calls to extract and analyze the financial data.")
            
            # Create the comprehensive commentary
            if all_commentary:
                # Join all commentary with proper paragraph formatting
                full_commentary = "\n\n".join(filter(None, all_commentary))
                
                # Clean and format the commentary for Excel cell
                formatted_commentary = clean_csv_value(full_commentary)
                
                # Create single general commentary CSV row
                csv_row = f'row{row_counter}: General Commentary,Document Summary,Overall Document Commentary,{formatted_commentary},N/A,Complete document analysis and extracted information summary'
                csv_output.append(csv_row)
                yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                row_counter += 1
            
            # Add cost summary if available
            cost_summary = result.get('cost_summary', {})
            if cost_summary:
                total_cost = cost_summary.get('total_cost_usd', 0)
                total_tokens = cost_summary.get('total_tokens', 0)
                api_calls = cost_summary.get('api_calls', 0)
                
                # Create CSV-formatted row for cost summary
                csv_row = f"row{row_counter}: Cost Summary,Processing Cost,Total LLM Cost,{clean_csv_value(f'${total_cost:.6f}')},N/A,{clean_csv_value(f'Tokens: {total_tokens:,} | API Calls: {api_calls}')}"
                csv_output.append(csv_row)
                yield f"data: {json.dumps({'type': 'row', 'content': csv_row})}\n\n"
                row_counter += 1
            
            # Send completion signal (no large CSV data to avoid JSON truncation)
            yield f"data: {json.dumps({'type': 'complete', 'total_rows': row_counter - 1, 'cost_summary': cost_summary})}\n\n"
            
        except Exception as e:
            yield f"data: {json.dumps({'status': 'error', 'error': str(e)})}\n\n"
    
    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Headers': 'Cache-Control',
        'X-Accel-Buffering': 'no'
    })

@app.route('/export_xlsx', methods=['POST'])
def export_xlsx():
    """Export CSV data to XLSX format with specific sheet name"""
    try:
        import io
        import csv
        from io import StringIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        csv_data = data.get('csv_data', '')
        
        if not csv_data:
            return jsonify({'error': 'No CSV data provided'}), 400
        
        # Create workbook with specific sheet name
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Extracted_data_comments"
        
        # Parse CSV data properly handling multi-line content
        if ws is not None:
            # Use proper CSV parsing to handle multi-line quoted fields
            csv_reader = csv.reader(StringIO(csv_data))
            
            for row_idx, row_line in enumerate(csv_reader, 1):
                # Check if this is a valid row (starts with "rowN:")
                if row_line and len(row_line) > 0 and row_line[0].startswith('row'):
                    # Extract actual data after "rowN: "
                    first_cell = row_line[0]
                    colon_idx = first_cell.find(': ')
                    if colon_idx != -1:
                        # Remove "rowN: " prefix from first column
                        row_line[0] = first_cell[colon_idx + 2:]
                        
                        # Add to worksheet
                        for col_idx, cell_value in enumerate(row_line, 1):
                            # Handle multi-line content in Excel cells
                            if isinstance(cell_value, str) and '\\n' in cell_value:
                                cell_value = cell_value.replace('\\n', '\n')
                            
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                            
                            # Enable text wrapping for multi-line content
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            # Style header row
                            if row_idx == 1:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
            # Auto-adjust column widths
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                print(f"Error adjusting column widths: {width_error}")
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='extracted_data_comments.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate XLSX: {str(e)}'}), 500

@app.route('/process', methods=['POST'])
def process():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    try:
        import pandas as pd
        
        # Process the structured JSON data using unified or legacy mode
        if USE_UNIFIED_CONTEXT_EXTRACTION:
            # Pass full structured data for comprehensive unified processing
            unified_result = process_structured_data_with_llm_unified(data)
            
            # Convert unified result to standard format
            result = convert_unified_to_standard_format(unified_result, data)
        else:
            # Use legacy processing with separate LLM calls and commentary matching
            result = process_structured_data_with_llm(data)
        
        # Use the enhanced data with context if available
        if 'enhanced_data_with_context' in result and result['enhanced_data_with_context']:
            clean_data = result['enhanced_data_with_context']
            
            # Add general commentary as a separate row if it exists
            if result.get('general_commentary'):
                clean_data.append({
                    'source': 'Document Text',
                    'type': 'General Commentary',
                    'field': 'Unmatched Commentary',
                    'value': result['general_commentary'][:500] + '...' if len(result['general_commentary']) > 500 else result['general_commentary'],
                    'page': 'N/A',
                    'context': '',
                    'has_context': False
                })
        # Fallback to enhanced_data_with_commentary for backward compatibility
        elif 'enhanced_data_with_commentary' in result and result['enhanced_data_with_commentary']:
            clean_data = result['enhanced_data_with_commentary']
            
            # Add context column if missing
            for row in clean_data:
                if 'context' not in row:
                    row['context'] = row.get('commentary', '')
                    row['has_context'] = bool(row['context'])
            
            # Add general commentary as a separate row if it exists
            if result.get('general_commentary'):
                clean_data.append({
                    'source': 'Document Text',
                    'type': 'General Commentary',
                    'field': 'Unmatched Commentary',
                    'value': result['general_commentary'][:500] + '...' if len(result['general_commentary']) > 500 else result['general_commentary'],
                    'page': 'N/A',
                    'context': '',
                    'has_context': False
                })
        else:
            # Fallback to original processing if enhanced data is not available
            df_data = []
            
            # Process tables
            if 'processed_tables' in result and result['processed_tables']:
                for i, table in enumerate(result['processed_tables']):
                    if table.get('structured_table') and not table['structured_table'].get('error'):
                        table_data = table['structured_table']
                        page = table.get('page', 'N/A')
                        
                        # Handle different table structures
                        if isinstance(table_data, dict):
                            for key, value in table_data.items():
                                if key != 'error':
                                    df_data.append({
                                        'source': f'Table {i+1}',
                                        'type': 'Table Data',
                                        'field': key,
                                        'value': str(value) if value else '',
                                        'page': page,
                                        'commentary': '',
                                        'has_commentary': False
                                    })
            
            # Process key-value pairs
            if 'processed_key_values' in result and result['processed_key_values']:
                kv_data = result['processed_key_values'].get('structured_key_values', {})
                if kv_data and not kv_data.get('error'):
                    for key, value in kv_data.items():
                        if key != 'error':
                            df_data.append({
                                'source': 'Key-Value Pairs',
                                'type': 'Structured Data',
                                'field': key,
                                'value': str(value) if value else '',
                                'page': 'N/A',
                                'commentary': '',
                                'has_commentary': False
                            })
            
            # Process document text with tabulation
            if 'processed_document_text' in result and result['processed_document_text']:
                for chunk_idx, chunk in enumerate(result['processed_document_text']):
                    # Handle tabulated document text structure
                    if 'table_headers' in chunk and 'table_rows' in chunk:
                        headers = chunk['table_headers']
                        rows = chunk['table_rows']
                        
                        # Add document text table structure
                        df_data.append({
                            'source': f'Document Text {chunk_idx+1}',
                            'type': 'Document Table',
                            'field': 'Headers',
                            'value': ' | '.join(headers),
                            'page': 'N/A',
                            'commentary': 'Tabulated document content',
                            'has_commentary': True,
                            'is_table_header': True,
                            'table_id': f'doc_{chunk_idx}',
                            'headers': headers,
                            'rows': rows
                        })
                        
                        # Add individual data points from document table
                        for row_idx, row in enumerate(rows):
                            for col_idx, cell_value in enumerate(row):
                                if col_idx < len(headers) and cell_value:
                                    df_data.append({
                                        'source': f'Document Text {chunk_idx+1}',
                                        'type': 'Document Data',
                                        'field': f'{headers[col_idx]}_Row_{row_idx+1}',
                                        'value': str(cell_value),
                                        'page': 'N/A',
                                        'commentary': '',
                                        'has_commentary': False,
                                        'table_id': f'doc_{chunk_idx}'
                                    })
                    
                    # Also handle extracted facts if available
                    if 'extracted_facts' in chunk and not chunk['extracted_facts'].get('error'):
                        facts = chunk['extracted_facts']
                        for key, value in facts.items():
                            if key != 'error' and value:
                                df_data.append({
                                    'source': f'Text Chunk {chunk_idx+1}',
                                    'type': 'Financial Data',
                                    'field': key,
                                    'value': str(value),
                                    'page': 'N/A',
                                    'commentary': '',
                                    'has_commentary': False
                                })
            
            # Clean and prepare data
            if df_data:
                # Filter out empty values directly from the list
                clean_data = [
                    item for item in df_data 
                    if item.get('value', '').strip() and item.get('value') != 'nan'
                ]
            else:
                clean_data = []
        
        # Return both original result and clean DataFrame data
        response = {
            **result,
            'dataframe_data': clean_data,
            'total_rows': len(clean_data)
        }
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/extract_structured', methods=['POST'])
def extract_structured():
    """Extract structured data from PDF using Amazon Textract"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are supported'}), 400
    
    try:
        # Extract structured data using Textract
        pdf_bytes = file.read()
        structured_data = extract_structured_data_from_pdf_bytes(pdf_bytes)
        
        return jsonify({
            'success': True,
            'structured_data': structured_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    data = request.json
    if not data or 'data' not in data:
        return jsonify({'error': 'No data provided'}), 400
    
    try:
        import pandas as pd
        df = pd.DataFrame(data['data'])
        pdf_bytes = export_to_pdf(df)
        
        return jsonify({
            'pdf': base64.b64encode(pdf_bytes).decode('utf-8')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
if __name
__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    print(f"🚀 Starting PDF Extraction Service on port {port}")
    print(f"📊 Modules loaded: {MODULES_LOADED}")
    print(f"🔧 Debug mode: {debug}")
    
    app.run(host='0.0.0.0', port=port, debug=debug)